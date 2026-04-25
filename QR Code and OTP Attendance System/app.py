from flask import Flask, render_template, request, jsonify, send_from_directory
import pandas as pd
import datetime
import os
import qrcode
import calendar
import smtplib
from email.message import EmailMessage

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

app = Flask(__name__)

# ================= CONFIG =================
TEACHER_USERNAME = "Admin"
TEACHER_PASSWORD = "admin@123"

SENDER_EMAIL = "gamercode669@gmail.com"
APP_PASSWORD = "fhioshanbbbihegf"

data = pd.read_excel("students.xlsx")
attendance = {}

GREEN = PatternFill(start_color="90EE90", fill_type="solid")
BLUE = PatternFill(start_color="87CEFA", fill_type="solid")
RED = PatternFill(start_color="FF9999", fill_type="solid")


# ================= ROUTES =================
@app.route("/")
def home():
    return render_template("index.html")


@app.route("/scanner")
def scanner():
    return render_template("scanner.html")


@app.route("/success")
def success():
    return render_template("success.html")


@app.route("/dashboard")
def login():
    return render_template("login.html")


@app.route("/dashboard-home")
def dashboard():
    return render_template("dashboard.html")


@app.route("/images/<path:filename>")
def images(filename):
    return send_from_directory("images", filename)


# ================= LOGIN =================
@app.route("/check-login", methods=["POST"])
def check_login():
    username = request.json.get("username")
    password = request.json.get("password")

    if username == TEACHER_USERNAME and password == TEACHER_PASSWORD:
        return jsonify({"status": "ok"})
    return jsonify({"status": "fail"})


# ================= QR EMAIL =================
@app.route("/generate-qr", methods=["POST"])
def generate_qr():
    try:
        roll = int(request.json["roll"])
    except:
        return jsonify({"status": "error", "msg": "Invalid Roll"})

    student = data[data["Roll No."] == roll]

    if student.empty:
        return jsonify({"status": "error", "msg": "Roll Not Found"})

    email = student.iloc[0]["Email"]

    qr_file = f"qr_{roll}.png"
    qrcode.make(f"Roll:{roll}").save(qr_file)

    try:
        msg = EmailMessage()
        msg["Subject"] = "Attendance QR Code"
        msg["From"] = SENDER_EMAIL
        msg["To"] = email
        msg.set_content("Scan this QR to mark attendance.")

        with open(qr_file, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="image",
                subtype="png",
                filename=qr_file
            )

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(SENDER_EMAIL, APP_PASSWORD)
            smtp.send_message(msg)

    except Exception as e:
        print("Email Error:", e)

    if os.path.exists(qr_file):
        os.remove(qr_file)

    return jsonify({"status": "sent"})


# ================= EXCEL =================
def update_excel(att):
    now = datetime.datetime.now()
    month_name = now.strftime("%B")
    year = now.year
    today = now.day

    file = f"{month_name}_Attendance.xlsx"
    total_students = len(data)

    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.title = month_name

        ws["A1"] = "Name"
        ws["A1"].font = Font(bold=True)

        days = calendar.monthrange(year, now.month)[1]

        for d in range(1, days + 1):
            col = d + 1
            cell = ws.cell(1, col)
            cell.value = f"{d:02d}-{month_name[:3]}"
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

            if calendar.weekday(year, now.month, d) == 6:
                cell.fill = RED

        for i in range(total_students):
            ws.cell(i + 2, 1).value = data.iloc[i]["Name"]

        ws.freeze_panes = "B2"
        wb.save(file)

    wb = load_workbook(file)
    ws = wb.active
    col = today + 1

    for i in range(total_students):
        row_no = i + 2
        name = ws.cell(row_no, 1).value

        status = None

        for roll, s in att.items():
            student = data[data["Roll No."] == roll]

            if not student.empty:
                if student.iloc[0]["Name"] == name:
                    status = s
                    break

        if status is None:
            continue

        cell = ws.cell(row_no, col)

        if status == "P":
            cell.value = "P"
            cell.fill = GREEN

        elif status == "Late":
            cell.value = "Late"
            cell.fill = BLUE

        elif status == "A":
            cell.value = "A"
            cell.fill = RED

        cell.alignment = Alignment(horizontal="center")

    wb.save(file)


# ================= AUTO ABSENT =================
def auto_mark_absent():
    now = datetime.datetime.now().time()

    if now < datetime.time(9, 30):
        return

    changed = False

    for _, row in data.iterrows():
        roll = row["Roll No."]

        if roll not in attendance:
            attendance[roll] = "A"
            changed = True

    if changed:
        update_excel(attendance)


# ================= MARK ATTENDANCE =================
@app.route("/mark-attendance", methods=["POST"])
def mark_attendance():
    try:
        roll = int(request.json["roll"])
    except:
        return jsonify({"status": "error", "msg": "Invalid Roll"})

    qr = request.json["qr"]
    teacher = request.json.get("teacher", False)

    if f"Roll:{roll}" not in qr:
        return jsonify({"status": "error", "msg": "Invalid QR"})

    if roll in attendance:
        return jsonify({"status": "error", "msg": "Already Marked"})

    student = data[data["Roll No."] == roll]

    if student.empty:
        return jsonify({"status": "error", "msg": "Student Not Found"})

    name = student.iloc[0]["Name"]
    current = datetime.datetime.now().time()

    if current < datetime.time(9, 5):
        status = "P"

    elif current <= datetime.time(9, 15):
        status = "Late"

    else:
        if teacher:
            status = "Late"
        else:
            return jsonify({"status": "permission"})

    attendance[roll] = status
    update_excel(attendance)

    return jsonify({
        "status": "success",
        "name": name,
        "roll": roll,
        "time": datetime.datetime.now().strftime("%H:%M:%S"),
        "att": status
    })


# ================= DASHBOARD =================
@app.route("/dashboard-data")
def dashboard_data():
    auto_mark_absent()

    file = datetime.datetime.now().strftime("%B") + "_Attendance.xlsx"

    if not os.path.exists(file):
        return jsonify({
            "today": {"present": 0, "late": 0, "absent": len(data)},
            "monthly": []
        })

    df = pd.read_excel(file)
    today_col = datetime.datetime.now().strftime("%d-%b")

    present = 0
    late = 0
    absent = 0

    if today_col in df.columns:
        col = df[today_col].fillna("").astype(str).str.strip()

        present = col.isin(["P", "Present", "present", "p"]).sum()
        late = col.isin(["Late", "late", "L", "l"]).sum()
        absent = col.isin(["A", "Absent", "absent", "a"]).sum()

    monthly = []
    date_cols = df.columns[1:]

    for _, row in df.iterrows():
        vals = row[date_cols].fillna("").astype(str).str.strip()

        p = vals.isin(["P", "Present", "present", "p"]).sum()
        l = vals.isin(["Late", "late", "L", "l"]).sum()
        a = vals.isin(["A", "Absent", "absent", "a"]).sum()

        total = p + l + a
        percent = round(((p + l) / total) * 100, 2) if total > 0 else 0

        monthly.append({
            "Name": row["Name"],
            "P": int(p),
            "L": int(l),
            "A": int(a),
            "Percent": percent
        })

    return jsonify({
        "today": {
            "present": int(present),
            "late": int(late),
            "absent": int(absent)
        },
        "monthly": monthly
    })


if __name__ == "__main__":
    app.run(debug=True)