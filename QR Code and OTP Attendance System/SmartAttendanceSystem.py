# -------- IMPORT LIBRARIES --------
import qrcode
import pandas as pd
import datetime
import random
import smtplib
from email.message import EmailMessage
import time
import os
import cv2
from pyzbar.pyzbar import decode

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import calendar


# -------- LOAD STUDENT DATABASE --------
data = pd.read_excel("students.xlsx")

attendance = []
scanned_rolls = set()

# -------- COLORS --------
green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
blue = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# -------- TEACHER PASSWORD --------
TEACHER_PASSWORD = "admin123"

# -------- FUNCTION: TIME STATUS --------
def get_status():
    now = datetime.datetime.now().time()

    if now < datetime.time(9,5):
        return "Present"
    elif now <= datetime.time(9,15):
        return "Late"
    else:
        return "Permission"


# -------- EMAIL CONFIG --------
sender_email = "gamercode669@gmail.com"
sender_password = "fhioshanbbbihegf"   # ⚠ Gmail App Password


# -------- STUDENT INPUT --------
roll = int(input("Enter your Roll Number: "))

student = data[data["Roll No."] == roll]

if student.empty:
    print("❌ Roll number not found!")
    exit()

name = student.iloc[0]["Name"]
email = student.iloc[0]["Email"]

print("✅ Student Found:", name)


# -------- GENERATE QR --------
now = datetime.datetime.now()
timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
session_id = random.randint(10000,99999)

qr_data = f"""
Attendance System
Roll:{roll}
Session:{session_id}
Time:{timestamp}
"""

qr_file = f"qr_{roll}.png"

qr = qrcode.make(qr_data)
qr.save(qr_file)

print("✅ QR Generated")


# -------- SEND EMAIL --------
msg = EmailMessage()
msg["Subject"] = "Attendance QR Code"
msg["From"] = sender_email
msg["To"] = email

msg.set_content(f"Hello {name}, scan this QR within 2 minutes.")

with open(qr_file, "rb") as f:
    msg.add_attachment(f.read(), maintype="image", subtype="png", filename=qr_file)

try:
    with smtplib.SMTP_SSL("smtp.gmail.com",465) as smtp:
        smtp.login(sender_email, sender_password)
        smtp.send_message(msg)
    print("📩 QR sent")

except Exception as e:
    print("❌ Email error:", e)


# -------- START SCANNER --------
cap = cv2.VideoCapture(0)
start_time = datetime.datetime.now()

print("📷 Camera Started... Scan QR")

while True:
    ret, frame = cap.read()

    if not ret:
        print("❌ Camera not working")
        break

    for qr in decode(frame):
        qr_data = qr.data.decode()

        if "Roll:" in qr_data:
            scanned_roll = int(qr_data.split("Roll:")[1].split("\n")[0])

            if scanned_roll not in scanned_rolls:
                scanned_rolls.add(scanned_roll)

                status = get_status()

                # -------- AFTER 9:15 --------
                if status == "Permission":
                    pwd = input("Enter Teacher Password: ")

                    if pwd == TEACHER_PASSWORD:
                        status = "Late"
                    else:
                        status = "Absent"

                student = data[data["Roll No."] == scanned_roll]

                if not student.empty:
                    name = student.iloc[0]["Name"]

                    attendance.append({
                        "Name": name,
                        "Status": status
                    })

                    print(f"✅ {name} → {status}")

    cv2.imshow("Scanner", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

    if (datetime.datetime.now() - start_time).seconds > 120:
        print("⏳ QR expired")
        break

cap.release()
cv2.destroyAllWindows()


# -------- DELETE QR --------
if os.path.exists(qr_file):
    os.remove(qr_file)


# -------- EXCEL REPORT SYSTEM (FINAL FORMAT) --------
file_name = "attendance_report.xlsx"

today = datetime.datetime.now()
day = today.day
month = today.month
year = today.year

# -------- CREATE FILE IF NOT EXISTS --------
if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1).value = "Name/Date"

    days_in_month = calendar.monthrange(year, month)[1]

    for d in range(1, days_in_month + 1):
        col = d + 1
        date_str = f"{d:02d}-{today.strftime('%b')}"
        ws.cell(row=1, column=col).value = date_str

        # Sunday red
        if datetime.date(year, month, d).weekday() == 6:
            ws.cell(row=1, column=col).fill = red

    # Add all students
    for i, row in data.iterrows():
        ws.cell(row=i+2, column=1).value = row["Name"]

    wb.save(file_name)

# -------- LOAD FILE --------
wb = load_workbook(file_name)
ws = wb.active

today_col = day + 1

# -------- MARK ATTENDANCE --------
for i in range(2, ws.max_row + 1):
    student_name = ws.cell(row=i, column=1).value

    status = "A"  # default

    for record in attendance:
        if record["Name"] == student_name:
            if record["Status"] == "Present":
                status = "P"
            elif record["Status"] == "Late":
                status = "Late"

    cell = ws.cell(row=i, column=today_col)
    cell.value = status

    if status == "P":
        cell.fill = green
    elif status == "Late":
        cell.fill = blue
    else:
        cell.fill = red

wb.save(file_name)

print("📊 Final attendance saved in professional format")